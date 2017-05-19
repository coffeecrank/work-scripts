import openpyxl
import http.client
import urllib.parse
import os
import warnings
import traceback

def main():
    warnings.simplefilter('ignore')

    wb_name, wb, ws = load_source_file()

    if wb_name != None:
        print('\nChecking chapter numbering...')
        ws = check_chapters(ws)

        print('Checking for non-HTML characters...')
        ws = check_characters(ws)

        print('\nSorting rows...')
        ws = row_sort(ws)

        print('\nGenerating XML files...')
        ws = create_xml_files(ws)

        wb.save(wb_name)

        input('\nAll done. Press ENTER to exit...')
    else:
        input('No flashcards found. Press ENTER to exit...')

def load_source_file():
    spreadsheets = []
    for f in os.listdir(os.getcwd()):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            if has_flashcards(f):
                spreadsheets.append(f)

    if len(spreadsheets) > 0:
        print('Flashcards source files:')
        i = 1
        for s in spreadsheets:
            print(str(i) + '. ' + s)
            i += 1
        running = True
        while running:
            number = input('\nChoose a number: ')
            try:
                wb_name = spreadsheets[int(number) - 1]
                wb = openpyxl.load_workbook(wb_name)
                for sheet in wb.get_sheet_names():
                    ws = wb.get_sheet_by_name(sheet)
                    chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)
                    if chapter_column != None and term_column != None and def_column != None:
                        break
                running = False
            except:
                print('\nYou entered a wrong number or an invalid character!\n')
    else:
        return None, None, None

    return wb_name, wb, ws

def has_flashcards(spreadsheet):
    wb = openpyxl.load_workbook(spreadsheet)
    
    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)
        if chapter_column != None and term_column != None and def_column != None and ws.cell(row = 2, column = 1).value != None:
            return True

    return False

def get_columns(ws):
    chapter_column = None
    term_column = None
    def_column = None
    deck_column = None
    url_column = None

    for c in range(1, ws.max_column + 1):
        if ws.cell(row = 1, column = c).value != None and type(ws.cell(row = 1, column = c).value) == str:
            if 'theme' in ws.cell(row = 1, column = c).value.lower() and deck_column == None:
                deck_column = c
            if 'chapter' in ws.cell(row = 1, column = c).value.lower() and 'theme' not in ws.cell(row = 1, column = c).value.lower() and chapter_column == None:
                chapter_column = c
            if 'term' in ws.cell(row = 1, column = c).value.lower() and term_column == None:
                term_column = c
            if 'definition' in ws.cell(row = 1, column = c).value.lower() and def_column == None:
                def_column = c
            if 'url' in ws.cell(row = 1, column = c).value.lower() and url_column == None:
                url_column = c

    return chapter_column, term_column, def_column, deck_column, url_column

def check_chapters(ws):
    chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)

    chapters = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row = r, column = chapter_column).value != None and not any(c == ws.cell(row = r, column = chapter_column).value for c in chapters):
            chapters.append(ws.cell(row = r, column = chapter_column).value)

    p_chapters = []
    normal_chapters = []
    for c in chapters:
        try:
            normal_chapters.append([c, int(c)])
        except:
            num = ''
            for char in c:
                if char.isdigit():
                    num += char
            p_chapters.append([c, int(num)])

    if len(p_chapters) > 0:
        normal_chapters = sorted(normal_chapters, key = lambda x: x[1])
        p_chapters = sorted(p_chapters, key = lambda x: x[1])
        
        new_p_chapters = []
        i = 0
        for c in p_chapters:
            new_p_chapters.append([c[0], i])
            i += 1

        new_normal_chapters = []
        for c in normal_chapters:
            new_normal_chapters.append([c[0], i])
            i += 1

        new_chapters = []
        for c in new_p_chapters:
            new_chapters.append(c)
        for c in new_normal_chapters:
            new_chapters.append(c)

        for r in range(2, ws.max_row + 1):
            if ws.cell(row = r, column = chapter_column).value != None:
                for c in new_chapters:
                    if ws.cell(row = r, column = chapter_column).value == c[0]:
                        ws.cell(row = r, column = chapter_column).value = c[1]

    return ws

def check_characters(ws):
    chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)

    for r in range(2, ws.max_row + 1):
        if ws.cell(row = r, column = term_column).value != None:
            term_val = ws.cell(row = r, column = term_column).value
            def_val = ws.cell(row = r, column = def_column).value

            if type(term_val) == str:
                while '  ' in term_val:
                    term_val = term_val.replace('  ', ' ')            
                while term_val[0] == ' ':
                    term_val = term_val.replace(' ', '', 1)          
                while term_val[len(term_val) - 1] == ' ':
                    term_val = term_val[:len(term_val) - 1]
                
                new_term_val = ''
                for char in term_val:
                    if char == '…':
                        new_term_val += '...'
                    elif char == '’':
                        new_term_val += '\''
                    elif char == '‘':
                        new_term_val += '\''
                    elif char == '”':
                        new_term_val += '\"'
                    elif char == '“':
                        new_term_val += '\"'
                    elif char == '\n':
                        new_term_val += ' '
                    elif char == '—':
                        new_term_val += '-'
                    elif char == '–':
                        new_term_val += '-'
                    else:
                        new_term_val += char
            else:
                new_term_val = term_val
            
            if type(def_val) == str:
                while '  ' in def_val:
                    def_val = def_val.replace('  ', ' ')            
                while def_val[0] == ' ':
                    def_val = def_val.replace(' ', '', 1)          
                while def_val[len(def_val) - 1] == ' ':
                    def_val = def_val[:len(def_val) - 1]
                
                new_def_val = ''
                for char in def_val:
                    if char == '…':
                        new_def_val += '...'
                    elif char == '’':
                        new_def_val += '\''
                    elif char == '‘':
                        new_def_val += '\''
                    elif char == '”':
                        new_def_val += '\"'
                    elif char == '“':
                        new_def_val += '\"'
                    elif char == '\n':
                        new_def_val += ' '
                    elif char == '—':
                        new_def_val += '-'
                    elif char == '–':
                        new_def_val += '-'
                    else:
                        new_def_val += char
            else:
                new_def_val = def_val

            ws.cell(row = r, column = term_column).value = new_term_val
            ws.cell(row = r, column = def_column).value = new_def_val
        
    return ws

def row_sort(ws):
    chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)
       
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row = r, column = chapter_column).value != None:
            rows.append([])
            for c in range(1, ws.max_column + 1):
                rows[r - 2].append(ws.cell(row = r, column = c).value)

    rows = sorted(rows, key = lambda x: int(x[chapter_column - 1]))

    for r in range(2, ws.max_row + 1):
        if ws.cell(row = r, column = chapter_column).value != None:
            for c in range(1, ws.max_column + 1):
                ws.cell(row = r, column = c).value = rows[r - 2][c - 1]
            
    return ws

def create_xml_files(ws):
    chapter_column, term_column, def_column, deck_column, url_column = get_columns(ws)

    chapters = []
    decks = []
    terms = []
    definitions = []
    urls = []

    yellow_color = openpyxl.styles.colors.YELLOW
    yellow_fill = openpyxl.styles.PatternFill(fill_type = 'solid', start_color = yellow_color, end_color = yellow_color)
    
    for r in range(2, ws.max_row + 1):
        if ws.cell(row = r, column = term_column).value != None:
            chapters.append(int(ws.cell(row = r, column = chapter_column).value))
            if deck_column != None:
                deck_name = ws.cell(row = r, column = deck_column).value
                if type(deck_name) == str:
                    while '\n' in deck_name:
                        deck_name = deck_name.replace('\n', '')
                    while ' ' in deck_name:
                        deck_name = deck_name.replace(' ', '')
                    decks.append(deck_name)
            if type(ws.cell(row = r, column = term_column).value) == str:
                term = ''
                for char in ws.cell(row = r, column = term_column).value:
                    if char == '&':
                        term += '&amp;'
                    else:
                        term += char
            else:
                term = ws.cell(row = r, column = term_column).value
            terms.append(term)
            if type(ws.cell(row = r, column = def_column).value) == str:
                definition = ''
                for char in ws.cell(row = r, column = def_column).value:
                    if char == '&':
                        definition += '&amp;'
                    else:
                        definition += char
            else:
                definition = ws.cell(row = r, column = def_column).value
            definitions.append(definition)
            if url_column != None:
                urls.append(ws.cell(row = r, column = url_column).value)
        
    if len(urls) > 0:
        while True:
            choice = input('\nDo you also want to check URL\'s? Enter "y" or "n": ')
            try:
                if choice.lower() == 'y':
                    print('\nChecking URL\'s...')
                    for i in range(len(urls)):
                        p = urllib.parse.urlparse(urls[i])
                        conn = http.client.HTTPSConnection(p.netloc)
                        conn.request('HEAD', p.path)
                        resp = conn.getresponse()
                        if resp.status >= 400:
                            ws.cell(row = i + 1, column = url_column).fill = yellow_fill
                    break
                elif choice.lower() == 'n':
                    break
            except:
                print('\nYour answer is not valid!')
            
        common_dirs = []
        for url in urls:
            dir_path = os.path.dirname(url)
            if dir_path[len(dir_path) - 1] != '/':
                dir_path += '/'
            common_dirs.append(dir_path)
        cutoff = os.path.commonprefix(common_dirs)

    if len(decks) > 0:
        cur_deck = 0
        start_point = 0
        for t in range(0, len(terms)):
            if t != 0 and ((chapters[t] != chapters[t - 1] or decks[t] != decks[t - 1]) or t == len(terms) - 1):
                if chapters[t] != chapters[t - 1] or t == len(terms) - 1:
                    cur_deck = 0
                name = 'workflow_'       
                if chapters[t - 1] < 10:
                    name += '0'
                name += str(chapters[t - 1])
                name += '_'
                if cur_deck < 10:
                    name += '0'
                name += str(cur_deck)
                cur_deck += 1
                name += '.xml'

                f = open(name, 'w', encoding = 'utf-8', errors = 'surrogateescape')
                template = open('template.xml', 'r', encoding = 'ascii', errors = 'surrogateescape')

                template_text = ''
                for line in template.readlines():
                    template_text += line
                    template_text += '\n'

                template.close()

                part_one = template_text[0:template_text.index('{CHAPTER}')]
                template_text = template_text[template_text.index('{CHAPTER}') + len('{CHAPTER}'):]
                part_two = template_text[0:template_text.index('{CARDS}')]
                template_text = template_text[template_text.index('{CARDS}') + len('{CARDS}'):]
                part_three = template_text[0:template_text.index('{CHAPTER}')]
                template_text = template_text[template_text.index('{CHAPTER}') + len('{CHAPTER}'):]
                part_four = template_text

                xml = part_one
                xml += str(chapters[t - 1])
                xml += part_two

                if len(urls) > 0:
                    for k in range(start_point, t):
                        xml += '<card><term audio=\"'
                        xml += urls[k].replace(cutoff, '')
                        xml += '\"><flashtext>'
                        xml += str(terms[k])
                        xml += '</flashtext></term><def><flashtext>'
                        xml += str(definitions[k])
                        xml += '</flashtext></def></card>\n'
                else:
                    for k in range(start_point, t):
                        xml += '<card><term><flashtext>'
                        xml += str(terms[k])
                        xml += '</flashtext></term><def><flashtext>'
                        xml += str(definitions[k])
                        xml += '</flashtext></def></card>\n'

                xml += part_three
                xml += str(chapters[t - 1])
                xml += part_four
                
                f.write(xml)
                f.close()

                start_point = t
    else:
        start_point = 0
        for t in range(0, len(terms)):
            if t != 0 and (chapters[t] != chapters[t - 1] or t == len(terms) - 1):
                if chapters[t] != chapters[t - 1] or t == len(terms) - 1:
                    cur_deck = 0
                name = 'workflow_'       
                if chapters[t - 1] < 10:
                    name += '0'
                name += str(chapters[t - 1])
                name += '_00.xml'

                f = open(name, 'w', encoding = 'utf-8', errors = 'surrogateescape')
                template = open('template.xml', 'r', encoding = 'ascii', errors = 'surrogateescape')

                template_text = ''
                for line in template.readlines():
                    template_text += line
                    template_text += '\n'

                template.close()

                part_one = template_text[0:template_text.index('{CHAPTER}')]
                template_text = template_text[template_text.index('{CHAPTER}') + len('{CHAPTER}'):]
                part_two = template_text[0:template_text.index('{CARDS}')]
                template_text = template_text[template_text.index('{CARDS}') + len('{CARDS}'):]
                part_three = template_text[0:template_text.index('{CHAPTER}')]
                template_text = template_text[template_text.index('{CHAPTER}') + len('{CHAPTER}'):]
                part_four = template_text

                xml = part_one
                xml += str(chapters[t - 1])
                xml += part_two

                if len(urls) > 0:
                    for k in range(start_point, t):
                        xml += '<card><term audio=\"'
                        xml += urls[k].replace(cutoff, '')
                        xml += '\"><flashtext>'
                        xml += str(terms[k])
                        xml += '</flashtext></term><def><flashtext>'
                        xml += str(definitions[k])
                        xml += '</flashtext></def></card>\n'
                else:
                    for k in range(start_point, t):
                        xml += '<card><term><flashtext>'
                        xml += str(terms[k])
                        xml += '</flashtext></term><def><flashtext>'
                        xml += str(definitions[k])
                        xml += '</flashtext></def></card>\n'

                xml += part_three
                xml += str(chapters[t - 1])
                xml += part_four
                
                f.write(xml)
                f.close()

                start_point = t

    return ws

if __name__ == '__main__':
    try:
        main()
    except:
        traceback.print_exc()
        input('\nPress ENTER to exit...')
