import openpyxl
import os
import warnings

class Cell:
    def __init__(self, cell):
        self.instance = cell

    def has_substring(self, substring):
        if type(self.instance.value) == str:
            if substring in self.instance.value.lower():
                return True
        return False

    def is_empty(self):
        if self.instance.value != None:
            if self.instance.value == '':
                return True
            return False
        return True

    def fix_spacing(self):
        if not self.is_empty() and type(self.instance.value) == str:
            while '  ' in self.instance.value:
                self.instance.value = self.instance.value.replace('  ', ' ')
            try:
                while self.instance.value[0] == ' ':
                    self.instance.value = self.instance.value[1:]
            except IndexError:
                pass
            try:
                while self.instance.value[-1] == ' ':
                    self.instance.value = self.instance.value[:-1]
            except IndexError:
                pass

    def fix_hidden_characters(self):
        if not self.is_empty():
            if repr(self.instance.value)[0] == '\'' and repr(self.instance.value)[-1] == '\'':
                self.instance.value = repr(self.instance.value)[1:-1]
            if type(self.instance.value) == str and '\\u200b' in self.instance.value:
                self.instance.value = self.instance.value.replace('\\u200b', '')
            if type(self.instance.value) == str and '\\n' in self.instance.value:
                self.instance.value = self.instance.value.replace('\\n', ' ')
            if type(self.instance.value) == str and '\\\'' in self.instance.value:
                self.instance.value = self.instance.value.replace('\\\'', '\'')

    def fix_cnow_characters(self):
        if not self.is_empty() and type(self.instance.value) == str:
            if '\"' in self.instance.value:
                self.instance.value = self.instance.value.replace('\"', '\'\'')
            if '“' in self.instance.value:
                self.instance.value = self.instance.value.replace('“', '\'\'')
            if '”' in self.instance.value:
                self.instance.value = self.instance.value.replace('”', '\'\'')
            if '″' in self.instance.value:
                self.instance.value = self.instance.value.replace('″', '\'\'')
            if '«' in self.instance.value:
                self.instance.value = self.instance.value.replace('«', '\'\'')
            if '»' in self.instance.value:
                self.instance.value = self.instance.value.replace('»', '\'\'')
            if '–' in self.instance.value:
                self.instance.value = self.instance.value.replace('–', '-')
            if '—' in self.instance.value:
                self.instance.value = self.instance.value.replace('—', '-')

class Workbook:
    def __init__(self):
        self.name = ''
        self.instance = None

    def load(self):
        while True:
            try:
                wb_name = input('Please copy the file name here (including .xlsx): ')
                wb_instance = openpyxl.load_workbook(wb_name)
                break
            except FileNotFoundError:
                print('\nWrong file name! Try again.\n')
        self.name = wb_name
        self.instance = wb_instance

    def save(self):
        self.instance.save(self.name)
        input('\nAll set! Press ENTER to exit...')

    def get_worksheets(self):
        ws_list = []
        for ws_name in self.instance.get_sheet_names():
            ws = self.instance.get_sheet_by_name(ws_name)
            ws_list.append([ws_name, ws])
        return ws_list

class Worksheet:
    def __init__(self, ws_name, ws):
        self.name = ws_name
        self.instance = ws
        self.asset_type_column = -1
        self.asset_description_column = -1

    def is_grid(self):
        for c in range(1, self.instance.max_column + 1):
            cell = Cell(self.instance.cell(row = 1, column = c))
            if cell.has_substring('asset description') or cell.has_substring('caption'):
                self.asset_description_column = c
            elif cell.has_substring('asset type'):
                self.asset_type_column = c
                return True
            elif cell.is_empty():
                break
        return False

    def fix(self):
        for r in range(1, self.instance.max_row + 1):
            asset_type_cell = Cell(self.instance.cell(row = r, column = self.asset_type_column))
            i = 0
            for c in range(1, self.instance.max_column + 1):
                cell = Cell(self.instance.cell(row = r, column = c))
                if cell.is_empty():
                    i += 1
                if i == 20:
                    break
                cell.fix_spacing()
                cell.fix_hidden_characters()
                if not asset_type_cell.is_empty():
                    if asset_type_cell.has_substring('cnow') and c != self.asset_type_column and c != self.asset_description_column:
                        cell.fix_cnow_characters()
                    
def main():
    warnings.simplefilter('ignore')
    
    workbook = Workbook()
    workbook.load()

    for ws_name, ws in workbook.get_worksheets():
        worksheet = Worksheet(ws_name, ws)
        if worksheet.is_grid():
            worksheet.fix()

    workbook.save()

if __name__ == '__main__':
    main()
